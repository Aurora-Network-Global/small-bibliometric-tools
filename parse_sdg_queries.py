import pandas_read_xml as pdx 
from collections import OrderedDict
import pandas as pd
import glob
from openpyxl import load_workbook
import xlsxwriter


# parse #text field from SDG query

def parse_query(sdg_xml):
    output=[]
    df=pdx.read_xml(sdg_xml, encoding="utf-8")
    query_keys=df.iloc[2,:]["aqd:query"]["aqd:query-definition"]
    for _ in range(len(query_keys)):
        if isinstance(query_keys[_]["aqd:query-lines"]["aqd:query-line"], list):
            for l in range(len(query_keys[_]["aqd:query-lines"]["aqd:query-line"])):
                output.append(query_keys[_]["aqd:query-lines"]["aqd:query-line"][l]["#text"])
        if isinstance(query_keys[_]["aqd:query-lines"]["aqd:query-line"], OrderedDict):
            output.append(query_keys[_]["aqd:query-lines"]["aqd:query-line"]["#text"])
    return output


# extract keywords from parsed text

def extract_keywords(parsed_query, unique=True):
    output=[]
    for _ in range(len(parsed_query)):
        output.append([x for x in parsed_query[_].split('"') if x not in parsed_query[_].split('"')[::2]])
    output=[x for y in output for x in y]
    if unique==True:
        output=list(set(output))
    return output


# extract all keywords per SDG query
# as argument insert a directory with downloaded SDG queries in .xml format, which can be found here: https://github.com/Aurora-Network-Global/sdg-queries

def extract_all_keywords(directory):
    output=[]
    files=glob.glob(directory)
    n=0
    for f in files:
        n+=1
        q=parse_query(f)
        k=extract_keywords(q)
        output.append(k)
    output=pd.DataFrame(output).transpose()
    return output


# RUN THE CODE:
# extract all keywords per SDG query

tab=extract_all_keywords(".../sdg-queries-master/*xml")


# create empty .xlsx file in given directory

book=xlsxwriter.Workbook(".../sdg_keywords.xlsx")


# set path with created empty .xlsx file

path=".../sdg_keywords.xlsx"
book=load_workbook(path)


# iniciate ExcelWriter

writer=pd.ExcelWriter(path, engine="openpyxl")
writer.book=book


# save SDG queries into created .xlx file
# new sheet is created per each SDG

for _ in range(17):
    tab.iloc[:,_].to_excel(writer, sheet_name="sdg"+str(_))

writer.save()
writer.close()
